import streamlit as st
import pandas as pd
from serpapi import GoogleSearch
from dotenv import load_dotenv
import os
import io
from openpyxl import load_workbook

# 環境変数の読み込み
load_dotenv()

# 複数のSerpAPIキーを取得（Streamlit Cloud対応）
def load_api_keys():
    """複数のSerpAPIキーを読み込む"""
    api_keys = []
    
    try:
        # Streamlit Cloudの場合はst.secretsから取得
        for i in range(1, 11):  # 最大10個のキーをサポート
            key_name = f"SERPAPI_KEY{i}" if i > 1 else "SERPAPI_KEY"
            key = st.secrets.get(key_name, None)
            if key and key != "your_serpapi_key_here":
                api_keys.append(key)
    except:
        # ローカル環境の場合は.envから取得
        for i in range(1, 11):  # 最大10個のキーをサポート
            key_name = f"SERPAPI_KEY{i}" if i > 1 else "SERPAPI_KEY"
            key = os.getenv(key_name)
            if key and key != "your_serpapi_key_here":
                api_keys.append(key)
    
    return api_keys

# APIキーリストを取得
API_KEYS = load_api_keys()

# セッションステートの初期化（現在使用中のキーのインデックス）
if 'current_api_key_index' not in st.session_state:
    st.session_state.current_api_key_index = 0
if 'failed_api_keys' not in st.session_state:
    st.session_state.failed_api_keys = set()

def get_current_api_key():
    """現在使用可能なAPIキーを取得"""
    if not API_KEYS:
        return None
    
    # 失敗したキーを除外して利用可能なキーを取得
    available_keys = [key for i, key in enumerate(API_KEYS) 
                      if i not in st.session_state.failed_api_keys]
    
    if not available_keys:
        return None
    
    # 現在のインデックスが範囲外なら0にリセット
    if st.session_state.current_api_key_index >= len(API_KEYS):
        st.session_state.current_api_key_index = 0
    
    return API_KEYS[st.session_state.current_api_key_index]

def switch_to_next_api_key():
    """次のAPIキーに切り替える"""
    # 現在のキーを失敗リストに追加
    st.session_state.failed_api_keys.add(st.session_state.current_api_key_index)
    
    # 次の利用可能なキーを探す
    for i in range(len(API_KEYS)):
        next_index = (st.session_state.current_api_key_index + 1 + i) % len(API_KEYS)
        if next_index not in st.session_state.failed_api_keys:
            st.session_state.current_api_key_index = next_index
            return True
    
    return False  # すべてのキーが失敗

def search_phone_number(store_name, prefecture=""):
    """
    SerpAPIを使用して店舗名と都道府県から電話番号を検索する
    複数のAPIキーに対応し、上限に達したら自動的に次のキーに切り替える
    
    Args:
        store_name (str): 検索する店舗名
        prefecture (str): 都道府県名（オプション）
        
    Returns:
        str: 見つかった電話番号、または見つからない場合は空文字列
    """
    if not API_KEYS:
        return "APIキー未設定"
    
    current_key = get_current_api_key()
    if not current_key:
        return "全てのAPIキーが上限に達しました"
    
    max_retries = len(API_KEYS)  # 最大リトライ回数 = キーの数
    
    for retry in range(max_retries):
        try:
            # 検索クエリを作成（店舗名 + 都道府県 + 電話番号）
            search_query = f"{store_name}"
            if prefecture and pd.notna(prefecture) and prefecture != "":
                search_query += f" {prefecture}"
            search_query += " 電話番号"
            
            # SerpAPIで検索
            params = {
                "engine": "google",
                "q": search_query,
                "api_key": current_key,
                "num": 5,
                "hl": "ja",
                "gl": "jp"
            }
            
            search = GoogleSearch(params)
            results = search.get_dict()
            
            # エラーチェック
            if "error" in results:
                error_message = results.get("error", "")
                # クォータエラーや認証エラーの場合は次のキーに切り替え
                if "quota" in error_message.lower() or "limit" in error_message.lower() or "credits" in error_message.lower():
                    if switch_to_next_api_key():
                        current_key = get_current_api_key()
                        continue  # 次のキーでリトライ
                    else:
                        return "全てのAPIキーが上限に達しました"
                else:
                    return f"APIエラー: {error_message}"
            
            # ナレッジグラフから電話番号を取得
            if "knowledge_graph" in results:
                kg = results["knowledge_graph"]
                if "phone" in kg:
                    return kg["phone"]
            
            # ローカルパックから電話番号を取得
            if "local_results" in results and len(results["local_results"]) > 0:
                local_result = results["local_results"][0]
                if "phone" in local_result:
                    return local_result["phone"]
            
            # オーガニック検索結果から電話番号を抽出（スニペット内）
            if "organic_results" in results:
                for result in results["organic_results"][:3]:
                    snippet = result.get("snippet", "")
                    # 簡易的な電話番号パターンマッチング
                    import re
                    phone_patterns = [
                        r'\d{2,4}-\d{2,4}-\d{4}',
                        r'\d{3}-\d{4}-\d{4}',
                        r'\d{10,11}'
                    ]
                    for pattern in phone_patterns:
                        match = re.search(pattern, snippet)
                        if match:
                            return match.group()
            
            return "見つかりませんでした"
            
        except Exception as e:
            error_str = str(e)
            # APIクォータエラーの場合は次のキーに切り替え
            if "quota" in error_str.lower() or "limit" in error_str.lower() or "429" in error_str:
                if switch_to_next_api_key():
                    current_key = get_current_api_key()
                    continue  # 次のキーでリトライ
                else:
                    return "全てのAPIキーが上限に達しました"
            else:
                return f"エラー: {error_str}"
    
    return "全てのAPIキーが上限に達しました"

def process_excel(uploaded_file, preview_only=False):
    """
    Excelファイルを処理し、店舗名から電話番号を検索してK列に追加する
    
    Args:
        uploaded_file: アップロードされたExcelファイル
        preview_only (bool): プレビューのみの場合True
        
    Returns:
        tuple: (処理済みDataFrame, 処理済みExcelファイル(bytes), 検索カウント, スキップカウント)
    """
    # Excelファイルを読み込み
    uploaded_file.seek(0)
    excel_data = pd.ExcelFile(uploaded_file)
    
    # 「架電リスト」シートを読み込み
    if "架電リスト" not in excel_data.sheet_names:
        st.error("「架電リスト」シートが見つかりません。")
        return None, None, 0, 0
    
    uploaded_file.seek(0)
    df = pd.read_excel(uploaded_file, sheet_name="架電リスト")
    
    # A列が「店舗名」であることを確認
    if df.columns[0] != "店舗名":
        st.warning(f"A列の列名が「店舗名」ではなく「{df.columns[0]}」です。処理を続行します。")
    
    # C列が都道府県であることを確認
    prefecture_col = df.columns[2] if len(df.columns) > 2 else None
    if prefecture_col and "都道府県" not in str(prefecture_col):
        st.info(f"C列の列名: 「{prefecture_col}」")
    
    # K列のインデックスは10（0始まり）
    # K列の列名を確認または作成
    if len(df.columns) < 11:
        # K列が存在しない場合は列を追加
        for i in range(len(df.columns), 11):
            df.insert(i, f'Unnamed_{i}', "")
    
    # K列（インデックス10）を「店舗番号」として設定
    col_k_name = df.columns[10] if len(df.columns) > 10 else '店舗番号'
    if col_k_name != '店舗番号' and 'Unnamed' not in str(col_k_name):
        # 既存の列名がある場合は保持
        pass
    else:
        df.rename(columns={col_k_name: '店舗番号'}, inplace=True)
    
    # プログレスバーを表示
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    # 各店舗名に対して電話番号を検索（K列が空の場合のみ）
    total_rows = len(df)
    search_count = 0
    skip_count = 0
    
    for idx, row in df.iterrows():
        store_name = row[df.columns[0]]  # A列の値
        prefecture = row[df.columns[2]] if len(df.columns) > 2 else ""  # C列の値（都道府県）
        current_phone = row.get('店舗番号', '')  # K列の現在の値
        
        # 店舗名が入力されており、かつK列（店舗番号）が空の場合のみ検索
        if pd.notna(store_name) and store_name != "":
            # K列が空（NaNまたは空文字列）の場合のみ検索
            if pd.isna(current_phone) or current_phone == "":
                search_count += 1
                search_text = f"{store_name} {prefecture}" if pd.notna(prefecture) else store_name
                status_text.text(f"検索中: {search_text} ({idx + 1}/{total_rows}) - 検索: {search_count}件, スキップ: {skip_count}件")
                phone_number = search_phone_number(str(store_name), str(prefecture) if pd.notna(prefecture) else "")
                df.at[idx, '店舗番号'] = phone_number
            else:
                skip_count += 1
                status_text.text(f"スキップ: {store_name} (既に電話番号あり) ({idx + 1}/{total_rows}) - 検索: {search_count}件, スキップ: {skip_count}件")
        
        # プログレスバーを更新
        progress_bar.progress((idx + 1) / total_rows)
    
    status_text.text(f"検索完了！検索: {search_count}件, スキップ: {skip_count}件")
    
    # 元のExcelファイルを読み込み、フォーマットと他のシートを保持
    uploaded_file.seek(0)
    wb = load_workbook(uploaded_file)
    
    # 「架電リスト」シートを取得
    ws = wb["架電リスト"]
    
    # K列（11列目）のデータのみ更新
    k_col_idx = 11  # Excelは1始まり
    for idx, row in df.iterrows():
        excel_row = idx + 2  # ヘッダー行を考慮（+1）、0始まりを1始まりに（+1）
        ws.cell(row=excel_row, column=k_col_idx, value=df.at[idx, '店舗番号'])
    
    # Excelファイルとして出力
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return df, output.getvalue(), search_count, skip_count

def main():
    st.set_page_config(
        page_title="店舗電話番号検索アプリ",
        page_icon="📞",
        layout="wide"
    )
    
    st.title("📞 店舗電話番号検索アプリ")
    st.markdown("---")
    
    # APIキーの確認
    if not API_KEYS:
        st.error("⚠️ SerpAPIキーが設定されていません。")
        
        with st.expander("🔑 設定方法"):
            st.markdown("""
            ### Streamlit Cloudの場合：
            1. アプリの「⋮」→「Settings」→「Secrets」タブ
            2. 以下の形式で入力：
            ```toml
            SERPAPI_KEY = "あなたのAPIキー1"
            SERPAPI_KEY2 = "あなたのAPIキー2"
            SERPAPI_KEY3 = "あなたのAPIキー3"
            ```
            
            ### ローカル環境の場合：
            1. `.env`ファイルに以下を記載：
            ```
            SERPAPI_KEY=あなたのAPIキー1
            SERPAPI_KEY2=あなたのAPIキー2
            SERPAPI_KEY3=あなたのAPIキー3
            ```
            
            ### APIキーの取得：
            1. https://serpapi.com/ でアカウントを作成
            2. APIキーを取得（複数のアカウントで複数のキーを取得可能）
            """)
        return
    
    # APIキーの状態を表示
    available_keys = len(API_KEYS) - len(st.session_state.failed_api_keys)
    
    if available_keys > 0:
        col1, col2 = st.columns([3, 1])
        with col1:
            st.success(f"✅ SerpAPIキー: {len(API_KEYS)}個設定済み")
        with col2:
            st.info(f"利用可能: {available_keys}/{len(API_KEYS)}")
    else:
        st.error(f"⚠️ 全てのAPIキー({len(API_KEYS)}個)が上限に達しています")
        st.info("新しいAPIキーを追加するか、翌月までお待ちください。")
    
    # 使い方の説明
    with st.expander("📖 使い方"):
        st.markdown("""
        1. 「架電リスト」シートを含むExcelファイルを用意してください
        2. A列に「店舗名」、C列に「都道府県」カラムがあることを確認してください
        3. ファイルをアップロードしてください
        4. 「電話番号を検索」ボタンをクリックしてください
        5. 検索結果のプレビューを確認してください
        6. 問題なければ、ダウンロードボタンから結果をダウンロードできます
        
        **重要**: 
        - 検索クエリ: 「店舗名 + 都道府県 + 電話番号」で検索します
        - K列の「店舗番号」カラムに検索結果の電話番号が記載されます
        - **K列に既にデータが入っている行はスキップされます**（既存データは保持）
        - 元のExcelファイルのフォーマット、他のシートもそのまま保持されます
        """)
    
    # ファイルアップロード
    st.subheader("📁 Excelファイルをアップロード")
    uploaded_file = st.file_uploader(
        "「架電リスト」シートを含むExcelファイルを選択してください",
        type=["xlsx", "xls"],
        help="Excelファイル（.xlsx または .xls）のみアップロード可能です"
    )
    
    # セッションステートの初期化
    if 'processed_df' not in st.session_state:
        st.session_state.processed_df = None
    if 'processed_file' not in st.session_state:
        st.session_state.processed_file = None
    if 'search_count' not in st.session_state:
        st.session_state.search_count = 0
    if 'skip_count' not in st.session_state:
        st.session_state.skip_count = 0
    
    if uploaded_file is not None:
        st.success(f"✅ ファイル「{uploaded_file.name}」がアップロードされました")
        
        # ファイルのプレビュー
        try:
            uploaded_file.seek(0)
            df_preview = pd.read_excel(uploaded_file, sheet_name="架電リスト", nrows=5)
            st.subheader("📋 元データプレビュー（最初の5行）")
            st.dataframe(df_preview, use_container_width=True)
            
            # ファイルポインタを先頭に戻す
            uploaded_file.seek(0)
        except Exception as e:
            st.error(f"プレビュー表示エラー: {str(e)}")
            return
        
        # 処理ボタン
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            if st.button("🔍 電話番号を検索", use_container_width=True, type="primary"):
                with st.spinner("電話番号を検索中..."):
                    uploaded_file.seek(0)
                    result_df, result_file, search_count, skip_count = process_excel(uploaded_file)
                    
                    if result_df is not None and result_file is not None:
                        # セッションステートに保存
                        st.session_state.processed_df = result_df
                        st.session_state.processed_file = result_file
                        st.session_state.search_count = search_count
                        st.session_state.skip_count = skip_count
                        st.session_state.uploaded_filename = uploaded_file.name
        
        # 検索結果のプレビューと統計情報
        if st.session_state.processed_df is not None:
            st.markdown("---")
            st.success("🎉 処理が完了しました！")
            
            # 統計情報を表示
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("検索実行", f"{st.session_state.search_count}件")
            with col2:
                st.metric("スキップ", f"{st.session_state.skip_count}件")
            with col3:
                total = st.session_state.search_count + st.session_state.skip_count
                st.metric("合計", f"{total}件")
            
            # 検索結果のプレビュー（更新された行のみ）
            st.subheader("📊 検索結果プレビュー")
            
            # K列に新しく追加されたデータを持つ行を抽出
            result_df = st.session_state.processed_df
            
            # 表示する列を選択（A列、C列、K列）
            display_cols = []
            if len(result_df.columns) > 0:
                display_cols.append(result_df.columns[0])  # A列（店舗名）
            if len(result_df.columns) > 2:
                display_cols.append(result_df.columns[2])  # C列（都道府県）
            if '店舗番号' in result_df.columns:
                display_cols.append('店舗番号')  # K列
            
            # 更新された行のみをフィルタリング
            updated_rows = result_df[result_df['店舗番号'].notna() & (result_df['店舗番号'] != "")]
            
            if len(updated_rows) > 0:
                st.dataframe(updated_rows[display_cols].head(20), use_container_width=True)
                if len(updated_rows) > 20:
                    st.info(f"プレビューは最初の20件のみ表示しています。全{len(updated_rows)}件が処理されました。")
            else:
                st.info("更新された行がありません。")
            
            # ダウンロードボタン
            st.subheader("💾 ファイルをダウンロード")
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                st.download_button(
                    label="📥 処理済みExcelファイルをダウンロード",
                    data=st.session_state.processed_file,
                    file_name=f"processed_{st.session_state.uploaded_filename}",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary"
                )
    
    # フッター
    st.markdown("---")
    st.markdown(
        "<div style='text-align: center; color: gray;'>Powered by SerpAPI & Streamlit</div>",
        unsafe_allow_html=True
    )

if __name__ == "__main__":
    main()


